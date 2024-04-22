from kivy.uix.screenmanager import Screen
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.scrollview import ScrollView

from kivy.core.window import Window

from jnius import autoclass
import os
from os.path import join
import openpyxl as op
from openpyxl.utils import get_column_letter

Environment = autoclass('android.os.Environment')
context = autoclass('org.kivy.android.PythonActivity').mActivity
files_dir = context.getFilesDir().getAbsolutePath()
file_path = os.path.join(files_dir, "password.xlsx")

wb = op.load_workbook(file_path)
ws = wb.active
column_max = ws.max_column

site_names = []
screens = []

for i in range(1, column_max):
    site_name = ws.cell(row=1, column=1 + i).value
    if site_name is not None:
        site_names.append(site_name)

root = GridLayout(cols=3, size_hint_y=None, height=0)
root.bind(minimum_height=root.setter("height"))
# root = BoxLayout(orientation="vertical")
scrollview = ScrollView(
    # size_hint=(1, None),
    do_scroll_x=False,
    size=(Window.width, Window.height),
)


class MainScreen(Screen):
    def __init__(self, **kwargs):
        super(MainScreen, self).__init__(**kwargs)
        self.name = "main"
        self.load_buttons()
        self.root = root
        self.scrollview = scrollview
        self.scrollview.add_widget(root)
        self.add_widget(scrollview)
        self.screens = screens

    def load_buttons(self):

        from utils.button import MyButton

        from utils.button import AddNewButton

        from utils.button import HelpButton

        root.add_widget(HelpButton(text="HELP", size_hint_y=None))

        root.add_widget(AddNewButton(text="Add New +", size_hint_y=None))

        for name in site_names:

            info_screen = self.create_info_screen(name)
            root.add_widget(
                MyButton(
                    text=name,
                    size_hint_y=None,
                )
            )
            screens.append(info_screen)

    def create_info_screen(self, name):

        wb = op.load_workbook(file_path)
        ws = wb.active

        info_screen = Screen(
            name=name
        )  # type()을 통한 동적 생성x => Screen클래스의 인스턴스를 생성하는 것으로 대신.

        # info_screen
        content_layout = BoxLayout(orientation="vertical")
        site_name_layout = BoxLayout(orientation="vertical")  # 최상단 layout

        site_name = Label(text=name)
        site_name_layout.add_widget(site_name)
        content_layout.add_widget(site_name_layout)

        info_id_layout = BoxLayout()

        info_id = Label(text="ID: ")
        info_id_layout.add_widget(info_id)
        content_layout.add_widget(info_id_layout)

        def find_column_by_value(ws, target_value):

            for col in ws.iter_cols():
                for cell in col:
                    if cell.value == target_value:
                        return get_column_letter(cell.column)

        target_value = name
        found_column = find_column_by_value(ws, target_value)

        if found_column:
            id_gonna_input = ws[f"{found_column}" + "2"].value
            input_id = Label(text=f"{id_gonna_input}")
            info_id_layout.add_widget(input_id)

        info_pw_layout = BoxLayout()

        info_pw = Label(text="PW: ")
        info_pw_layout.add_widget(info_pw)
        content_layout.add_widget(info_pw_layout)

        if found_column:
            pw_gonna_input = ws[f"{found_column}" + "3"].value
            input_pw = Label(text=f"{pw_gonna_input}")
            info_pw_layout.add_widget(input_pw)

        button_layout = BoxLayout()

        from utils.button import CopyButton

        btn_copy_id = CopyButton(
            text="copy id",
            text_to_copy=id_gonna_input,
        )
        btn_copy_pw = CopyButton(
            text="copy pw",
            text_to_copy=pw_gonna_input,
        )
        button_layout.add_widget(btn_copy_id)
        button_layout.add_widget(btn_copy_pw)
        from utils.button import DeleteButton

        button_layout.add_widget(DeleteButton(text="Delete"))
        content_layout.add_widget(button_layout)

        info_screen.add_widget(content_layout)

        from utils.button import ReturnButton

        return_button_layout = BoxLayout()
        return_button = ReturnButton(text="Return")
        return_button_layout.add_widget(return_button)

        content_layout.add_widget(return_button_layout)

        return info_screen
