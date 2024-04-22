from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.popup import Popup

from kivy.core.clipboard import Clipboard

from jnius import autoclass
import os
from os.path import join
import openpyxl as op
from openpyxl.utils import get_column_letter

Environment = autoclass('android.os.Environment')
context = autoclass('org.kivy.android.PythonActivity').mActivity
files_dir = context.getFilesDir().getAbsolutePath()
file_path = os.path.join(files_dir, "password.xlsx")


class MyButton(Button):
    def __init__(self, **kwargs):
        super(MyButton, self).__init__(**kwargs)
        self.bind(on_press=self.switch_to_info_screen)

    def switch_to_info_screen(self, instance):

        sm = instance.parent.parent.parent.parent
        sm.current = self.text


class ReturnButton(Button):
    def __init__(self, **kwargs):
        super(ReturnButton, self).__init__(**kwargs)
        self.bind(on_press=self.return_to_main_screen)

    def return_to_main_screen(self, instance):
        sm = instance.parent.parent.parent.parent
        sm.current = "main"


class AddNewButton(Button):
    def __init__(self, **kwargs):
        super(AddNewButton, self).__init__(**kwargs)
        self.bind(on_press=self.add_new_info)

    def add_new_info(self, instance):
        sm = instance.parent.parent.parent.parent
        sm.current = "new"


class SaveButton(Button):
    def __init__(self, **kwargs):
        super(SaveButton, self).__init__(**kwargs)
        self.bind(on_press=self.save_infos)

    def save_infos(self, instance):

        wb = op.load_workbook(file_path)
        ws = wb.active
        column_max = ws.max_column
        datasets = []

        add_new_screen = instance.parent.parent.parent
        input_site_name = add_new_screen.site_name_input.text
        input_id = add_new_screen.id_input.text
        input_pw = add_new_screen.pw_input.text

        if input_site_name.startswith("[") or input_site_name.endswith("]"):
            from utils.notification import BracketNotification

            popup_bracket = Popup(
                title="Notification",
                size_hint=(0.7, 0.3),
            )

            notification = BracketNotification(
                popup_bracket,
                text="Don't start with '[' and don't end with']'\n(The brackets used at deleting logic.)",
            )
            popup_bracket.content = notification
            popup_bracket.open()
        else:
            datasets.append(input_site_name)
            datasets.append(input_id)
            datasets.append(input_pw)

            if "" not in datasets:
                if column_max == 1:
                    for i, dataset in enumerate(datasets, start=1):
                        ws.cell(row=i, column=2, value=dataset)
                elif column_max >= 2:
                    for i, dataset in enumerate(datasets, start=1):
                        ws.cell(row=i, column=column_max + 1, value=dataset)

                """saving"""
                wb.save(file_path)

                sm = instance.parent.parent.parent.parent
                main_screen = sm.get_screen("main")

                from utils.button import MyButton

                main_screen.root.add_widget(
                    MyButton(text=input_site_name, size_hint_y=None)
                )
                wb.save(file_path)

                screen_added = main_screen.create_info_screen(input_site_name)
                sm.add_widget(screen_added)

                add_new_screen.site_name_input.text = ""
                add_new_screen.id_input.text = ""
                add_new_screen.pw_input.text = ""

                sm.current = "main"

            else:
                from utils.notification import ValidateNotification

                popup_validate = Popup(
                    title="Notification",
                    size_hint=(0.7, 0.3),
                )

                notification = ValidateNotification(
                    popup_validate,
                    text="You should not add the 'Blank'.\nIf you wanna add only id or pw,\njust input like 'sample'.",
                )

                popup_validate.content = notification
                popup_validate.open()


class CopyButton(Button):
    def __init__(self, text_to_copy, **kwargs):
        super(CopyButton, self).__init__(**kwargs)
        self.bind(on_press=self.copy_text)
        self.text_to_copy = text_to_copy

    def copy_text(self, instance):
        text_to_copy = self.text_to_copy
        Clipboard.copy(text_to_copy)

        from utils.notification import CopyNotification

        popup = Popup(
            title="Notification",
            size_hint=(0.7, 0.3),
        )
        notification = CopyNotification(popup, text="Copied!")
        popup.content = notification
        popup.open()


class DeleteButton(Button):
    def __init__(self, **kwargs):
        super(DeleteButton, self).__init__(**kwargs)
        self.bind(on_press=self.open_popup)

    def open_popup(self, instance):
        from kivy.uix.popup import Popup

        content = BoxLayout(orientation="vertical")
        content.add_widget(Label(text="Are you sure for deleting?"))

        btn_del_txt = instance.parent.parent.parent.name
        delete_sure_btn = DeleteSureButton(
            text=f"Delete [{btn_del_txt}]",
            screen_manager=instance.parent.parent.parent.parent,
            button_instance=instance.parent.parent.parent.parent,
        )
        content.add_widget(delete_sure_btn)
        cancel_btn = Button(text="Cancel")

        content.add_widget(cancel_btn)

        popup = Popup(
            title="Popup for deleting",
            content=content,
            size_hint=(0.7, 0.3),
            auto_dismiss=False,
        )

        delete_sure_btn.bind(on_press=popup.dismiss)
        cancel_btn.bind(on_press=popup.dismiss)
        popup.open()


class DeleteSureButton(Button):
    def __init__(self, screen_manager, button_instance, **kwargs):
        super(DeleteSureButton, self).__init__(**kwargs)
        self.screen_manager = screen_manager
        self.button_instance = button_instance
        self.bind(on_press=self.real_delete)

    def real_delete(self, instance):

        wb = op.load_workbook(file_path)
        ws = wb.active

        import re

        text = self.text
        match = re.search(r"\[([^]]+)\]", text)

        if match:
            text_modified = match.group(1)

        def find_column_by_value(ws, target_value):

            for col in ws.iter_cols():
                for cell in col:
                    if cell.value == target_value:
                        return get_column_letter(cell.column)

        target_value = text_modified
        found_column = find_column_by_value(ws, target_value)

        screen_manager = self.screen_manager
        for screen in screen_manager.screens:
            if screen.name == "main":
                for i, widget in enumerate(screen.walk()):
                    if i == 2:
                        buttons = widget.children
                        for button in buttons:
                            if button.text == text_modified:
                                main_screen = screen_manager.get_screen("main")
                                main_screen.root.remove_widget(button)

        if found_column:
            for i in range(1, 4):
                ws[f"{found_column}" + f"{i}"].value = None

            wb.save(file_path)

        screen_delete = screen_manager.get_screen(f"{text_modified}")
        screen_manager.remove_widget(screen_delete)

        screen_manager.current = "main"


class HelpButton(Button):
    def __init__(self, **kwargs):
        super(HelpButton, self).__init__(**kwargs)
        self.bind(on_press=self.switch_to_help_screen)

    def switch_to_help_screen(self, instance):
        sm = instance.parent.parent.parent.parent
        sm.current = "help"
