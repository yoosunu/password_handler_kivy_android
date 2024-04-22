import kivy

kivy.require("2.3.0")

from kivy.app import App
from kivy.uix.screenmanager import ScreenManager, FadeTransition

from utils.new_screen import AddNewScreen
from utils.help_screen import HelpScreen


from jnius import autoclass
import os
from os.path import join
import openpyxl as op


Environment = autoclass('android.os.Environment')
context = autoclass('org.kivy.android.PythonActivity').mActivity
files_dir = context.getFilesDir().getAbsolutePath()
file_path = os.path.join(files_dir, "password.xlsx")

def workbook_create(file_path):
    wb = op.Workbook()
    ws = wb.active

    BASES = ["SITE", "ID", "PW"]
    for i, BASE in enumerate(BASES, start=1):
        ws.cell(row=i, column=1, value=BASE)

    wb.save(file_path)
    wb.close()
    
if not os.path.exists(file_path):
    workbook_create(file_path)
    
else:
    pass
    
wb = op.load_workbook(file_path)
ws = wb.active
column_max = ws.max_column
site_names = []

for i in range(1, column_max):
    site_name = ws.cell(row=1, column=1 + i).value
    if site_name is not None:
        site_names.append(site_name)
        

class PasswordApp(App):

    def build(self):

        sm = ScreenManager(transition=FadeTransition())
        
        from utils.main_screen import MainScreen
        screen_main = MainScreen()
        screens = screen_main.screens

        for screen in screens:
            sm.add_widget(screen)

        screen_add_new = AddNewScreen()
        help_screen = HelpScreen()

        sm.add_widget(screen_main)
        sm.add_widget(screen_add_new)
        sm.add_widget(help_screen)

        sm.current = "main"
	
        return sm

    def on_stop(self):
        wb = op.load_workbook(file_path)
        wb.save(file_path)
        wb.close()


if __name__ == "__main__":
    PasswordApp().run()
